from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import mysql.connector
from datetime import datetime
from openpyxl import Workbook, load_workbook
import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Image, Spacer
)
from reportlab.lib.enums import TA_CENTER
import json
from fpdf import FPDF
from flask_mail import Mail, Message
import os


app = Flask(__name__)
app.secret_key = "zxcvbnm951"

db = mysql.connector.connect(
    host="localhost",
    user="root",
    password="root",
    database="payroll"
)

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'mohammedyousuf432003@gmail.com'
app.config['MAIL_PASSWORD'] = 'gweo mxde oqqv vlbi'

mail = Mail(app)


# ---------- STAFF CODE GENERATOR ----------
def generate_staff_code(cursor, department, date_of_join):
    dept_codes = {
        "ENGLISH": "01", "TAMIL": "02", "ARABIC": "03", "URDU": "04",
        "HINDI": "05", "FRENCH": "06", "HISTORICAL STUDIES": "07",
        "ECONOMICS": "08", "SOCIOLOGY": "09", "COMMERCE": "10",
        "CORPORATE SECRETARYSHIP": "11", "MATHEMATICS": "12",
        "PHYSICS": "13", "CHEMISTRY": "14", "BOTANY": "15",
        "ZOOLOGY": "16", "COMPUTER SCIENCE": "17",
        "COMPUTER APPLICATION": "18", "INFORMATION SYSTEM MANAGEMENT": "19",
        "BUSINESS ADMIN": "20", "BANK MANAGEMENT": "21",
        "BIOTECHNOLOGY": "22", "INFORMATION TECHNOLOGY": "23",
        "ACCOUNTS & FINANCE": "24",
        "CRIMINOLOGY & POLICE ADMINISTRATION": "25",
        "DEFENCE & STRATEGIC STUDIES": "26",
        "ELECTRONIC MEDIA": "27", "PROFESSIONAL ACCOUNTING": "28",
        "ARTIFICIAL INTELLIGENCE": "29", "DATASCIENCE": "30", "PHYSICAL EDUCATION": "31",
        "LIBRARY": "32"
    }

    dept_code = dept_codes[department.upper()]
    year_code = str(date_of_join.year)[-2:]

    cursor.execute(
        "SELECT COUNT(*) AS series FROM employees WHERE staff_code LIKE %s",
        (f"{year_code}{dept_code}%",)
    )
    row = cursor.fetchone()
    series = (row["series"] or 0) + 1

    return f"{year_code}{dept_code}{series:03}"

# ---------- INDEX ----------
@app.route("/")
def index():
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM employees")
    employees = cursor.fetchall()
    cursor.close()
    return render_template("index.html", employees=employees)

# ---------- ADD EMPLOYEE ----------
@app.route("/add_employee", methods=["GET", "POST"])
def add_employee():
    if request.method == "POST":
        cursor = db.cursor(dictionary=True)
        try:
            doj_str = request.form["date_of_join"]
            doj = datetime.strptime(doj_str, "%Y-%m-%d").date()

            staff_code = generate_staff_code(
                cursor,
                request.form["department"],
                doj
            )

            values = (
                staff_code,
                request.form["name"],
                request.form["department"],
                request.form["designation"],
                request.form["category"],
                request.form.get("aadhar"),
                request.form.get("pan"),
                request.form.get("bank_account"),
                request.form.get("pf_account"),
                request.form.get("basic", 0),
                request.form.get("hra", 0),
                request.form.get("da", 0),
                request.form.get("cca", 0),
                request.form.get("ir", 0),
                request.form.get("ma", 0),
                request.form.get("special_allowance", 0),
                doj_str,
                request.form.get("dob"),
                request.form.get("esi", 0),
                request.form.get("insurance", 0),
                request.form.get("pf", 0),
                request.form.get("professional_tax", 0),
                request.form.get("teachers_guild", 0),
                request.form.get("ntsw", 0),
                request.form.get("icrs", 0),
                request.form.get("ncswp", 0),
                request.form.get("nta", 0),
                request.form.get("gross_salary", 0),
                request.form.get("total_deductions", 0),
                request.form.get("net_salary", 0),
                request.form.get("phone"),
                request.form.get("email"),
                request.form.get("increment_month"),
            )

            cursor.execute("""
                INSERT INTO employees
                (staff_code, name, department, designation, category,
                 aadhar, pan, bank_account, pf_account,
                 basic, hra, da, cca, ir, ma, special_allowance,
                 date_of_join, dob,
                 esi, insurance, pf, professional_tax, teachers_guild,
                 ntsw, icrs, ncswp, nta,
                 gross_salary, total_deductions, net_salary,
                 phone, email, increment_month)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,
                        %s,%s,%s,%s,%s,%s,%s,
                        %s,%s,
                        %s,%s,%s,%s,%s,
                        %s,%s,%s,%s,
                        %s,%s,%s,%s,%s,%s)
            """, values)

            db.commit()
            flash("Employee Added Successfully ✅", "success")

        except Exception as e:
            flash(str(e), "error")
        finally:
            cursor.close()

        return redirect(url_for("add_employee"))

    return render_template("add_employee.html")

# ---------- DOWNLOAD EMPLOYEE TEMPLATE ----------
@app.route("/download_employee_template")
def download_employee_template():
    cursor = db.cursor()
    cursor.execute("DESCRIBE employees")

    columns = [
        r[0] for r in cursor.fetchall()
        if r[0] not in ("id", "staff_code", "created_at", "updated_at")
    ]
    cursor.close()

    wb = Workbook()
    ws = wb.active
    ws.append(columns)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="employee_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ---------- UPLOAD EMPLOYEE EXCEL ----------
@app.route("/upload_employee_excel", methods=["POST"])
def upload_employee_excel():
    file = request.files.get("excel_file")
    if not file:
        flash("No file selected ❌", "error")
        return redirect(url_for("index"))

    wb = load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    headers = [h for h in rows[0] if h]
    data_rows = rows[1:]

    cursor = db.cursor(dictionary=True)

    for row in data_rows:
        if all(v is None for v in row):
            continue

        data = dict(zip(headers, row))
        doj = data["date_of_join"]
        if isinstance(doj, str):
            doj = datetime.strptime(doj, "%Y-%m-%d").date()

        staff_code = generate_staff_code(cursor, data["department"], doj)

        cols = ["staff_code"] + list(data.keys())
        vals = [staff_code] + list(data.values())

        placeholders = ",".join(["%s"] * len(vals))
        cursor.execute(
            f"INSERT INTO employees ({','.join(cols)}) VALUES ({placeholders})",
            vals
        )

    db.commit()
    cursor.close()
    flash("Employees uploaded successfully ✅", "success")
    return redirect(url_for("index"))

# ---------- DELETE EMPLOYEE ----------
# @app.route("/delete_employee/<staff_code>", methods=["POST"])
# def delete_employee(staff_code):
#     cursor = db.cursor()
#     try:
#         cursor.execute(
#             "DELETE FROM employees WHERE staff_code = %s",
#             (staff_code,)
#         )
#         db.commit()
#         flash(f"Employee {staff_code} deleted successfully 🗑️", "success")
#     except Exception as e:
#         flash(str(e), "error")
#     finally:
#         cursor.close()

#     return redirect(url_for("index"))

@app.route("/delete_employee/<staff_code>", methods=["POST"])
def delete_employee(staff_code):
    delete_type = request.form.get("delete_type")
    delete_reason = request.form.get("delete_reason")

    cursor = db.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT staff_code, name, department, designation, date_of_join
            FROM employees WHERE staff_code=%s
        """, (staff_code,))
        emp = cursor.fetchone()

        if not emp:
            flash("Employee not found", "error")
            return redirect(url_for("index"))

        now = datetime.now()

        cursor.execute("""
            INSERT INTO deleted_employees
            (staff_code, name, department, designation, date_of_join,
             delete_type, delete_reason, deleted_at, deleted_month, deleted_year)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            emp["staff_code"],
            emp["name"],
            emp["department"],
            emp["designation"],
            emp["date_of_join"],
            delete_type,
            delete_reason if delete_type == "Other" else delete_type,
            now,
            now.strftime("%B"),
            now.year
        ))

        cursor.execute(
            "DELETE FROM employees WHERE staff_code=%s",
            (staff_code,)
        )

        db.commit()
        flash("Employee deleted successfully 🗑️", "success")

    except Exception as e:
        db.rollback()
        flash(str(e), "error")

    finally:
        cursor.close()

    return redirect(url_for("index"))

# ---------- SEARCH EMPLOYEE ----------
@app.route('/search_employee')
def search_employee():
    query = request.args.get('query', '')
    cursor = db.cursor(dictionary=True)
    sql = """
    SELECT staff_code, name, department, designation
    FROM employees
    WHERE staff_code LIKE %s
       OR name LIKE %s
       OR department LIKE %s
       OR designation LIKE %s
    """
    like_query = f"%{query}%"
    cursor.execute(sql, (like_query, like_query, like_query, like_query))
    results = cursor.fetchall()
    cursor.close()
    return jsonify(results)

# ---------- GET EMPLOYEE ----------
@app.route("/get_employee/<staff_code>")
def get_employee(staff_code):
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        "SELECT * FROM employees WHERE staff_code=%s",
        (staff_code,)
    )
    emp = cursor.fetchone()
    cursor.close()
    return jsonify(emp)

@app.route("/update_employee_modal", methods=["POST"])
def update_employee_modal():
    d = request.json
    cursor = db.cursor()

    cursor.execute("""
        UPDATE employees SET
            name=%s,
            department=%s,
            designation=%s,
            category=%s,
            aadhar=%s,
            pan=%s,
            bank_account=%s,
            pf_account=%s,

            basic=%s,
            hra=%s,
            da=%s,
            cca=%s,
            ir=%s,
            ma=%s,
            special_allowance=%s,

            esi=%s,
            insurance=%s,
            pf=%s,
            professional_tax=%s,
            teachers_guild=%s,
            ntsw=%s,
            icrs=%s,
            ncswp=%s,
            nta=%s
        WHERE staff_code=%s
    """, (
        d["name"], d["department"], d["designation"], d["category"],
        d["aadhar"], d["pan"], d["bank_account"], d["pf_account"],
        d["basic"], d["hra"], d["da"], d["cca"], d["ir"], d["ma"], d["special_allowance"],
        d["esi"], d["insurance"], d["pf"], d["professional_tax"],
        d["teachers_guild"], d["ntsw"], d["icrs"], d["ncswp"], d["nta"],
        d["staff_code"]
    ))

    db.commit()
    cursor.close()
    return jsonify({"message": "Employee updated successfully ✅"})

# ---------- SELECTED DELETE ----------
# @app.route("/bulk_delete_employee", methods=["POST"])
# def bulk_delete_employee():
#     data = request.get_json()
#     staff_codes = data.get("staff_codes")

#     if not staff_codes:
#         return jsonify({"message": "No employees selected"}), 400

#     cursor = db.cursor()
#     format_strings = ",".join(["%s"] * len(staff_codes))

#     cursor.execute(
#         f"DELETE FROM employees WHERE staff_code IN ({format_strings})",
#         tuple(staff_codes)
#     )

#     db.commit()
#     cursor.close()

#     return jsonify({"message": "Selected employees deleted successfully ✅"})

@app.route("/bulk_delete_employee", methods=["POST"])
def bulk_delete_employee():
    delete_type = request.form.get("delete_type")
    delete_reason = request.form.get("delete_reason")
    staff_codes = request.form.get("bulk_staff_codes")

    if not staff_codes:
        return jsonify({"message": "No employees selected"}), 400

    staff_codes = json.loads(staff_codes)

    cursor = db.cursor(dictionary=True)

    try:
        format_strings = ",".join(["%s"] * len(staff_codes))

        cursor.execute(f"""
            SELECT staff_code, name, department, designation, date_of_join
            FROM employees
            WHERE staff_code IN ({format_strings})
        """, tuple(staff_codes))

        employees = cursor.fetchall()
        now = datetime.now()

        for emp in employees:
            cursor.execute("""
                INSERT INTO deleted_employees
                (staff_code, name, department, designation, date_of_join,
                 delete_type, delete_reason, deleted_at, deleted_month, deleted_year)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                emp["staff_code"],
                emp["name"],
                emp["department"],
                emp["designation"],
                emp["date_of_join"],
                delete_type,
                delete_reason if delete_type == "Other" else delete_type,
                now,
                now.strftime("%B"),
                now.year
            ))

        cursor.execute(
            f"DELETE FROM employees WHERE staff_code IN ({format_strings})",
            tuple(staff_codes)
        )

        db.commit()
        flash(f"{len(staff_codes)} employees deleted successfully 🗑️", "success")

    except Exception as e:
        db.rollback()
        return jsonify({"message": str(e)}), 500

    finally:
        cursor.close()

    return redirect(url_for("index"))

# ---------- GENERATE PAYBILL ----------
@app.route("/generate_paybill/<month>/<int:year>")
def generate_paybill(month, year):

    cursor = db.cursor(dictionary=True)

    table_name = f"paybill_{year}_{month.lower()}"

    allowed_months = [
        "january","february","march","april","may","june",
        "july","august","september","october","november","december"
    ]

    if month.lower() not in allowed_months:
        return "Invalid month", 400


    # 🔒 CHECK ALREADY GENERATED
    cursor.execute("""
        SELECT * FROM paybill_register
        WHERE pay_month=%s AND pay_year=%s
    """, (month, year))

    if cursor.fetchone():
        return f"❌ Paybill already generated for {month} {year}", 400

    # 🔥 CREATE PAYBILL TABLE
    cursor.execute(f"""
    CREATE TABLE {table_name} (
    id INT AUTO_INCREMENT PRIMARY KEY,
    staff_code INT,
    name VARCHAR(100),
    department VARCHAR(100),
    designation VARCHAR(100),

    basic DECIMAL(10,2),
    da DECIMAL(10,2),
    hra DECIMAL(10,2),
    cca DECIMAL(10,2),
    ir DECIMAL(10,2),
    ma DECIMAL(10,2),
    special_allowance DECIMAL(10,2),

    esi DECIMAL(10,2),
    pf DECIMAL(10,2),
    professional_tax DECIMAL(10,2),
    insurance DECIMAL(10,2),

    gross_salary DECIMAL(10,2),
    total_deductions DECIMAL(10,2),
    net_salary DECIMAL(10,2)
    )
    """)


    # 📥 COPY DATA
    cursor.execute(f"""
    INSERT INTO {table_name}
    (
    staff_code, name, department, designation,
    basic, da, hra, cca, ir, ma, special_allowance,
    esi, pf, professional_tax, insurance,
    gross_salary, total_deductions, net_salary
    )
    SELECT
    staff_code, name, department, designation,
    basic, da, hra, cca, ir, ma, special_allowance,
    esi, pf, professional_tax, insurance,
    gross_salary, total_deductions, net_salary
    FROM employees
    """)


    # 📝 REGISTER ENTRY
    cursor.execute("""
        INSERT INTO paybill_register
        (pay_month, pay_year, table_name)
        VALUES (%s,%s,%s)
    """, (month, year, table_name))

    db.commit()  # 🔥 VERY IMPORTANT

    # 📤 FETCH STORED DATA
    cursor.execute(f"SELECT * FROM {table_name}")
    data = cursor.fetchall()

    # 🔽 GENERATE PDF & RETURN
    return generate_paybill_pdf(data, month, year)

# ---------- GENERATE PAYBILL PDF ----------
def generate_paybill_pdf(data, month, year):

    buffer = io.BytesIO()

    pdf = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    elements = []

    # ---------------- HEADER ----------------
    try:
        logo = Image("static/collegelogo.png", 50, 50)
    except:
        logo = ""

    header = Table([
        [
            logo,
            Paragraph(
                "<b>THE NEW COLLEGE (AUTONOMOUS)</b><br/>ROYAPETTAH, CHENNAI – 14",
                ParagraphStyle(
                    "h",
                    alignment=TA_CENTER,
                    fontSize=14
                )
            ),
            ""
        ]
    ], colWidths=[60, 650, 60])

    header.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))

    elements.append(header)

    elements.append(
        Paragraph(
            f"<b>PAY BILL – {month.upper()} {year}</b>",
            ParagraphStyle(
                "t",
                alignment=TA_CENTER,
                fontSize=12,
                spaceAfter=10
            )
        )
    )

    # ---------------- TABLE HEADER ----------------
    table_data = [[
        "Staff",
        "Name",
        "Basic", "DA", "HRA", "CCA", "IR", "MA", "Spl",
        "ESI", "PF", "PT", "Ins",
        "Gross", "Deduction", "Net"
    ]]

    # ---------------- TOTALS ----------------
    totals = {
        "basic": 0, "da": 0, "hra": 0, "cca": 0,
        "ir": 0, "ma": 0, "spl": 0,
        "esi": 0, "pf": 0, "pt": 0, "ins": 0,
        "gross": 0, "ded": 0, "net": 0
    }

    for r in data:
        table_data.append([
            r["staff_code"],
            r["name"],
            r["basic"],
            r["da"],
            r["hra"],
            r["cca"],
            r["ir"],
            r["ma"],
            r["special_allowance"],
            r["esi"],
            r["pf"],
            r["professional_tax"],
            r["insurance"],
            r["gross_salary"],
            r["total_deductions"],
            r["net_salary"]
        ])

        totals["basic"] += r["basic"] or 0
        totals["da"] += r["da"] or 0
        totals["hra"] += r["hra"] or 0
        totals["cca"] += r["cca"] or 0
        totals["ir"] += r["ir"] or 0 
        totals["ma"] += r["ma"] or 0
        totals["spl"] += r["special_allowance"] or 0
        totals["esi"] += r["esi"] or 0 
        totals["pf"] += r["pf"] or 0
        totals["pt"] += r["professional_tax"] or 0
        totals["ins"] += r["insurance"] or 0
        totals["gross"] += r["gross_salary"] or 0
        totals["ded"] += r["total_deductions"] or 0
        totals["net"] += r["net_salary"] or 0

    # ---------------- TOTAL ROW ----------------
    table_data.append([
        "",
        "TOTAL",
        totals["basic"], totals["da"], totals["hra"],
        totals["cca"], totals["ir"], totals["ma"], totals["spl"],
        totals["esi"], totals["pf"], totals["pt"], totals["ins"],
        totals["gross"], totals["ded"], totals["net"]
    ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            45, 85,
            45, 45, 45, 45, 40, 40, 45,
            45, 45, 45, 45,
            50, 50, 50
        ]
    )

    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (2,1), (-1,-1), "RIGHT"),
        ("FONT", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#E6E6E6")),
        ("FONTSIZE", (0,0), (-1,-1), 8),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    # ---------------- SIGN ----------------
    sign = Table([
        ["Prepared By", "", "THE NEW COLLEGE"],
        ["", "", ""]
    ], colWidths=[300, 200, 300])

    sign.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING", (0,0), (-1,-1), 20),
    ]))

    elements.append(sign)

    pdf.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"PayBill_{month}_{year}.pdf",
        mimetype="application/pdf"
    )

# ---------- PAYBILL STATUS ----------
@app.route("/paybill_status")
def paybill_status():
    cursor = db.cursor(dictionary=True)
    cursor.execute("""
        SELECT pay_month, pay_year
        FROM paybill_register
    """)
    data = cursor.fetchall()
    return jsonify(data)

# DOWNLOAD BANK STATEMENT (EXCEL)
@app.route("/download_bank_statement/<month>/<int:year>")
def download_bank_statement(month, year):

    cursor = db.cursor(dictionary=True)

    # 🔍 Step 1: Get paybill table name from register
    cursor.execute("""
        SELECT table_name
        FROM paybill_register
        WHERE pay_month = %s AND pay_year = %s
    """, (month, year))

    row = cursor.fetchone()
    if not row:
        cursor.close()
        return "Paybill not generated for this month ❌", 400

    table_name = row["table_name"]

    # 🔗 Step 2: JOIN paybill table with employees table
    query = f"""
        SELECT
            p.staff_code,
            e.name,
            e.bank_account,
            p.net_salary
        FROM {table_name} p
        JOIN employees e
            ON p.staff_code = e.staff_code
        ORDER BY p.staff_code
    """

    cursor.execute(query)
    data = cursor.fetchall()
    cursor.close()

    if not data:
        return "No data found ❌", 400

    # 📊 Step 3: Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}_{year}"

    # Header row
    ws.append([
        "Staff Code",
        "Employee Name",
        "Bank Account",
        "Net Salary"
    ])

    # Data rows
    for r in data:
        ws.append([
            r["staff_code"],
            r["name"],
            r["bank_account"] or "",
            float(r["net_salary"])
        ])

    # 💾 Step 4: Save Excel to memory
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    # 📤 Step 5: Send file
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"Bank_Statement_{month}_{year}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# DOWNLOAD BANK STATEMENT (PDF)
@app.route("/download_bank_statement_pdf/<month>/<int:year>")
def download_bank_statement_pdf(month, year):

    cursor = db.cursor(dictionary=True)

    # 🔍 Get paybill table name
    cursor.execute("""
        SELECT table_name
        FROM paybill_register
        WHERE pay_month=%s AND pay_year=%s
    """, (month, year))

    row = cursor.fetchone()
    if not row:
        cursor.close()
        return "Paybill not generated ❌", 400

    table_name = row["table_name"]

    # 🔗 JOIN paybill + employees
    cursor.execute(f"""
        SELECT
            p.staff_code,
            e.name,
            e.bank_account,
            p.net_salary
        FROM {table_name} p
        JOIN employees e
            ON p.staff_code = e.staff_code
        ORDER BY p.staff_code
    """)

    data = cursor.fetchall()
    cursor.close()

    if not data:
        return "No data found ❌", 400

    # 📄 Create PDF in memory
    buffer = io.BytesIO()

    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    elements = []

    # 📊 Table Data
    table_data = [
        ["Staff Code", "Employee Name", "Bank Account", "Net Salary"]
    ]

    total = 0
    for r in data:
        table_data.append([
            r["staff_code"],
            r["name"],
            r["bank_account"] or "",
            f"{float(r['net_salary']):,.2f}"
        ])
        total += float(r["net_salary"])

    # 🧮 Total row
    table_data.append([
        "", "TOTAL", "", f"{total:,.2f}"
    ])

    # 📐 Create table
    table = Table(table_data, colWidths=[70, 160, 140, 90])

    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (3,1), (3,-1), "RIGHT"),
        ("FONT", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("BACKGROUND", (0,-1), (-1,-1), colors.whitesmoke),
    ]))

    elements.append(table)

    pdf.build(elements)
    buffer.seek(0)

    # 📤 Send PDF
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Bank_Statement_{month}_{year}.pdf",
        mimetype="application/pdf"
    )

# ---------------- GENERATE PAYSLIPS (ALL STAFF) ----------------
@app.route("/generate_payslips")
def generate_payslips():
    cursor = db.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            staff_code,
            name,
            department,
            designation,
            basic,
            (hra + da + cca + ir + ma + special_allowance) AS allowance,
            (esi + pf + professional_tax + insurance) AS deduction,
            net_salary
        FROM employees
    """)

    data = cursor.fetchall()
    cursor.close()

    return jsonify(data)

# ---------------- SENDING PAYSLIP ( INDIVIDUAL ) ----------------
@app.route("/send_payslip/<staff_code>")
def send_payslip(staff_code):
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM employees WHERE staff_code=%s", (staff_code,))
    emp = cursor.fetchone()
    cursor.close()

    if not emp:
        return "Employee not found", 404

    # Generate PDF (reuse preview logic or separate function)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=11)

    pdf.cell(200, 10, "PAYSLIP", ln=True, align="C")
    pdf.ln(5)
    pdf.cell(200, 8, f"Staff Code : {emp['staff_code']}", ln=True)
    pdf.cell(200, 8, f"Name       : {emp['name']}", ln=True)
    pdf.cell(200, 8, f"Net Salary : Rs. {emp['net_salary']}", ln=True)

    file_path = f"Payslip_{staff_code}.pdf"
    pdf.output(file_path)

    msg = Message(
        subject="Your Monthly Payslip",
        sender="mohammedyousuf432003@gmail.com",
        recipients=[emp["email"]]
    )
    msg.body = f"""
        Dear {emp['name']},

        Please find attached your payslip.

        Regards,
        Payroll Team
        """
    with open(file_path, "rb") as f:
        msg.attach(file_path, "application/pdf", f.read())

    mail.send(msg)
    return "Payslip sent successfully"

@app.route("/preview_payslip_pdf/<staff_code>")
def preview_payslip_pdf(staff_code):
    cursor = db.cursor(dictionary=True)
    cursor.execute("""
        SELECT
            staff_code,
            name,
            department,
            designation,
            basic,
            (hra + da + cca + ir + ma + special_allowance) AS allowance,
            (esi + pf + professional_tax + insurance) AS deduction,
            net_salary
        FROM employees
        WHERE staff_code=%s
    """, (staff_code,))
    emp = cursor.fetchone()
    cursor.close()

    if not emp:
        return "Employee not found", 404

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=11)

    pdf.cell(200, 10, "PAYSLIP", ln=True, align="C")
    pdf.ln(5)
    pdf.cell(200, 8, f"Staff Code : {emp['staff_code']}", ln=True)
    pdf.cell(200, 8, f"Name       : {emp['name']}", ln=True)
    pdf.cell(200, 8, f"Department : {emp['department']}", ln=True)
    pdf.cell(200, 8, f"Designation: {emp['designation']}", ln=True)

    pdf.ln(5)
    pdf.cell(200, 8, f"Basic      : Rs. {emp['basic']}", ln=True)
    pdf.cell(200, 8, f"Allowance  : Rs. {emp['allowance']}", ln=True)
    pdf.cell(200, 8, f"Deduction  : Rs. {emp['deduction']}", ln=True)
    pdf.ln(3)
    pdf.cell(200, 10, f"NET SALARY : Rs. {emp['net_salary']}", ln=True)

    file_name = f"Payslip_{staff_code}.pdf"
    pdf.output(file_name)

    return send_file(file_name, mimetype="application/pdf", as_attachment=False)

@app.route("/download_all_payslips")
def download_all_payslips():
    cursor = db.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            staff_code, name, department, designation,
            basic,
            (hra + da + cca + ir + ma + special_allowance) AS allowance,
            (esi + pf + professional_tax + insurance) AS deduction,
            net_salary
        FROM employees
    """)
    employees = cursor.fetchall()
    cursor.close()

    if not employees:
        return "No employees found", 404

    pdf = FPDF()
    pdf.set_font("Arial", size=11)

    for emp in employees:
        pdf.add_page()

        pdf.cell(200, 10, "PAYSLIP", ln=True, align="C")
        pdf.ln(5)

        pdf.cell(200, 8, f"Staff Code : {emp['staff_code']}", ln=True)
        pdf.cell(200, 8, f"Name       : {emp['name']}", ln=True)
        pdf.cell(200, 8, f"Department : {emp['department']}", ln=True)
        pdf.cell(200, 8, f"Designation: {emp['designation']}", ln=True)

        pdf.ln(5)
        pdf.cell(200, 8, f"Basic      : Rs. {emp['basic']}", ln=True)
        pdf.cell(200, 8, f"Allowance  : Rs. {emp['allowance']}", ln=True)
        pdf.cell(200, 8, f"Deduction  : Rs. {emp['deduction']}", ln=True)

        pdf.ln(3)
        pdf.cell(200, 10, f"NET SALARY : Rs. {emp['net_salary']}", ln=True)

    file_name = "All_Employees_Payslips.pdf"
    pdf.output(file_name)

    return send_file(file_name, as_attachment=True)

# @app.route("/send_bulk_payslips", methods=["POST"])
# def send_bulk_payslips():
#     try:
#         cursor = db.cursor(dictionary=True)
#         cursor.execute("""
#             SELECT staff_code, name, email, net_salary
#             FROM employees
#             WHERE email IS NOT NULL
#         """)
#         employees = cursor.fetchall()
#         cursor.close()

#         if not employees:
#             return jsonify({"message": "No employees with email"}), 400

#         for emp in employees:
#             msg = Message(
#                 subject="Monthly Payslip",
#                 sender=app.config["MAIL_USERNAME"],
#                 recipients=[emp["email"]]
#             )

#             msg.body = f"""
#                 Dear {emp['name']},

#                 Your monthly payslip has been generated.

#                 Net Salary : Rs. {emp['net_salary']}

#                 Regards,
#                 Payroll Team
#                             """

#             mail.send(msg)

#         return jsonify({"message": "Payslips sent to all employees successfully ✅"})

#     except Exception as e:
#         print("MAIL ERROR:", e)
#         return jsonify({"message": str(e)}), 500

@app.route("/send_bulk_payslips", methods=["POST"])
def send_bulk_payslips():
    try:
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                staff_code,
                name,
                email,
                department,
                designation,
                basic,
                (hra + da + cca + ir + ma + special_allowance) AS allowance,
                (esi + pf + professional_tax + insurance) AS deduction,
                net_salary
            FROM employees
            WHERE email IS NOT NULL
        """)
        employees = cursor.fetchall()
        cursor.close()

        if not employees:
            return jsonify({"message": "No employees with email"}), 400

        sent_count = 0

        for emp in employees:
            pdf_path = create_payslip_pdf(emp)

            msg = Message(
                subject="Monthly Payslip",
                sender=app.config["MAIL_USERNAME"],
                recipients=[emp["email"]]
            )

            msg.body = f"""
                Dear {emp['name']},

                Please find attached your monthly payslip.

                Net Salary : Rs. {emp['net_salary']}

                Regards,
                Payroll Team
                            """

            with open(pdf_path, "rb") as f:
                msg.attach(
                    filename=os.path.basename(pdf_path),
                    content_type="application/pdf",
                    data=f.read()
                )

            mail.send(msg)
            sent_count += 1

            os.remove(pdf_path)  # 🔥 cleanup

        return jsonify({
            "message": f"Payslips sent successfully to {sent_count} employees ✅"
        })

    except Exception as e:
        print("MAIL ERROR:", e)
        return jsonify({"message": str(e)}), 500


def create_payslip_pdf(emp):
    folder = "temp_payslips"
    os.makedirs(folder, exist_ok=True)

    file_path = f"{folder}/Payslip_{emp['staff_code']}.pdf"

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=11)

    pdf.cell(200, 10, "PAYSLIP", ln=True, align="C")
    pdf.ln(5)

    pdf.cell(200, 8, f"Staff Code : {emp['staff_code']}", ln=True)
    pdf.cell(200, 8, f"Name       : {emp['name']}", ln=True)
    pdf.cell(200, 8, f"Department : {emp['department']}", ln=True)
    pdf.cell(200, 8, f"Designation: {emp['designation']}", ln=True)

    pdf.ln(5)
    pdf.cell(200, 8, f"Basic      : Rs. {emp['basic']}", ln=True)
    pdf.cell(200, 8, f"Allowance  : Rs. {emp['allowance']}", ln=True)
    pdf.cell(200, 8, f"Deduction  : Rs. {emp['deduction']}", ln=True)

    pdf.ln(3)
    pdf.cell(200, 10, f"NET SALARY : Rs. {emp['net_salary']}", ln=True)

    pdf.output(file_path)

    return file_path


if __name__ == "__main__":
    app.run(debug=True)

